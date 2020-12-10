import xml.etree.ElementTree as ET
import uuid
import re
import openpyxl

def apply_indent(elem, level = 0):
    # tab = space * 2
    indent = "\n" + level * "  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = indent + "  "
        if not elem.tail or not elem.tail.strip():
            elem.tail = indent
        for elem in elem:
            apply_indent(elem, level+1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = indent
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = indent

class ArxmlParser:
    def __init__(self, filepath, excelpath):
        ReadExcel = openpyxl.load_workbook(excelpath, data_only=True)  # pure data read except calculator of excel cell
        self.wb = ReadExcel['CAN IF List_v0.1']
        self.max_row = self.wb.max_row
        self.max_column = self.wb.max_column
        for col in self.wb.columns: # get excel column address
            #get column number which need for code
            if col[3].value == "CAN \nChannel": #it have to include escape charactor(\n, space)
                self.CANChannelCol = col[3].column
            elif col[3].value == "Rx/Tx":
                self.DirCol = col[3].column
            elif col[3].value == "CAN message":
                self.MsgCol = col[3].column
            elif col[3].value == "Length\n(Bit)":
                self.LengthCol = col[3].column
            elif col[3].value == "Cycle Time":
                self.CycleCol = col[3].column
            elif col[3].value == "Sender":
                self.SenderCol = col[3].column
            elif col[3].value == "Signal Name":
                self.SigCol = col[3].column
            elif col[2].value == "Variable Name":
                self.VarNameCol = col[2].column
            elif col[2].value == "Variable Type":
                self.TypeCol = col[2].column
            elif col[2].value == "Variable 초기값":
                self.InitCol = col[2].column
            elif col[2].value == "Time-out Flag":
                self.TimoutCol = col[2].column
            elif col[2].value == "Invalid Flag":
                self.InvalCol = col[2].column

        with open(filepath, 'r') as f:
            xml_string = f.read()
            xml_string = re.sub('<AUTOSAR (.*?)+>', '<AUTOSAR>', xml_string, count=1) # it is better to xml parsing after namespase was removed
            self.root = ET.fromstring(xml_string) # change to xml from 'string'

    def CCAN_extract_data(self):
        message = list()
        Message_Ind = list()
        Timeout_Ind = list()

        m = re.compile('var_cdcu.+')

        for i in range(5, self.max_row):
            CANChannel = self.wb.cell(row=i, column=self.CANChannelCol).value
            RxTx = self.wb.cell(row=i, column=self.DirCol).value
            if CANChannel == "C-CANFD" and RxTx == "Rx":
                Message = self.wb.cell(row=i, column=self.MsgCol).value
                Cycle = self.wb.cell(row=i, column=self.CycleCol).value
                Variable = self.wb.cell(row=i, column=self.VarNameCol).value
                p = m.match(str(Variable)) # except for unnecessary message
                if p:
                    message.append([Message, int(Cycle)])
                else:
                    pass

        # remove duplicated factor through 'set'
        message = set(map(tuple, message))  # remove 2 dimension list
        # it should be changed list type, which is able to access
        message = list(message)

        for msg, cycle in message:
            Message_Ind.append("Rt_{Message}_Ind".format(Message=msg))
            if cycle > 0:
                Message_Ind.append("RTout_{Message}_Ind".format(Message=msg))
            else:
                pass

        return Message_Ind

    def PCAN_extract_data(self):
        message = list()
        Message_Ind = list()

        m = re.compile('var_cdcu.+')

        for i in range(5, self.max_row):
            CANChannel = self.wb.cell(row=i, column=self.CANChannelCol).value
            RxTx = self.wb.cell(row=i, column=self.DirCol).value
            if CANChannel == "P-CANFD" and RxTx == "Rx":
                Message = self.wb.cell(row=i, column=self.MsgCol).value
                Cycle = self.wb.cell(row=i, column=self.CycleCol).value
                Variable = self.wb.cell(row=i, column=self.VarNameCol).value
                p = m.match(str(Variable)) # except for unnecessary message
                if p:
                    message.append([Message, int(Cycle)])
                else:
                    pass

        # remove duplicated factor through 'set'
        message = set(map(tuple, message))  # remove 2 dimension list
        # it should be changed list type, which is able to access
        message = list(message)

        for msg, cycle in message:
            Message_Ind.append("Rt_{Message}_Ind".format(Message=msg))
            if cycle > 0:
                Message_Ind.append("RTout_{Message}_Ind".format(Message=msg))
            else:
                pass

        return Message_Ind

    def GCAN_extract_data(self):
        message = list()
        Message_Ind = list()

        m = re.compile('var_cdcu.+')

        for i in range(5, self.max_row):
            CANChannel = self.wb.cell(row=i, column=self.CANChannelCol).value
            RxTx = self.wb.cell(row=i, column=self.DirCol).value
            if CANChannel == "G-CANFD" and RxTx == "Rx":
                Message = self.wb.cell(row=i, column=self.MsgCol).value
                Cycle = self.wb.cell(row=i, column=self.CycleCol).value
                Variable = self.wb.cell(row=i, column=self.VarNameCol).value
                p = m.match(str(Variable)) # except for unnecessary message
                if p:
                    message.append([Message, int(Cycle)])
                else:
                    pass

        # remove duplicated factor through 'set'
        message = set(map(tuple, message))  # remove 2 dimension list
        # it should be changed list type, which is able to access
        message = list(message)

        for msg, cycle in message:
            Message_Ind.append("Rt_{Message}_Ind".format(Message=msg)) # message ind runnable
            if cycle > 0:
                Message_Ind.append("RTout_{Message}_Ind".format(Message=msg)) # timeout ind runnable
            else:
                pass

        return Message_Ind

    def xml_gen(self, message):
        # search place where it is inserted runnable tag
        tag = self.root.find('./AR-PACKAGES/AR-PACKAGE/ELEMENTS/APPLICATION-SW-COMPONENT-TYPE/INTERNAL-BEHAVIORS/SWC-INTERNAL-BEHAVIOR/RUNNABLES')

        #make xmltree about runnable
        for i in message:
            node1 = ET.Element("RUNNABLE-ENTITY")
            node1.attrib['UUID'] = uuid.uuid1().__str__() # uuid random generate
            node2 = ET.SubElement(node1, 'SHORT-NAME')
            node2.text = i
            node3 = ET.SubElement(node1, 'MINIMUM-START-INTERVAL')
            node3.text = "0.0"
            node4 = ET.SubElement(node1, 'SYMBOL')
            node4.text = i
            #ET.dump(node1) # preview xml
            tag.append(node1)

    def make_xml(self, filepath):
        apply_indent(self.root)
        str_root = ET.tostring(self.root).decode() # change to 'string' from xml, if python version is 3 over it should need to decode() to change 'str
        namespace = '<AUTOSAR xmlns="http://autosar.org/schema/r4.0" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:schemaLocation="http://autosar.org/schema/r4.0 autosar_4-0-3.xsd">'
        str_root = re.sub('<AUTOSAR>', namespace, str_root, count=1) # add to namespace at 'AUTOSAR' tag

        self.root = ET.fromstring(str_root)  # change to xml from 'string'
        ET.register_namespace('','http://autosar.org/schema/r4.0')  # This must be necessary in order to remove 'ns:0' when xml is created
        self.tree = ET.ElementTree(self.root)  # create xml
        self.tree.write(filepath, encoding="UTF-8", xml_declaration=True)  # generate xml
