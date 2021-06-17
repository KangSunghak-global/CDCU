import openpyxl
import re
import codeStyle
import os

class CANCodeGenerate:

    def __init__(self, excelpath):
        ReadExcel = openpyxl.load_workbook(excelpath, data_only=True) # pure data read except calculator of excel cell
        self.wb = ReadExcel['CAN IF List_v2.0']
        self.max_row = self.wb.max_row
        self.max_column = self.wb.max_column
        for col in self.wb.columns: # get excel column address
            #get column number which need for code
            if col[3].value == "CAN \nChannel": #it have to include escape charactor(\n, space)
                self.CANChannelCol = col[3].column
            elif col[3].value == "Rx/Tx":
                self.DirCol = col[3].column
            elif col[3].value == "CAN message":
                self.MsgCol = col[3].column
            elif col[3].value == "Length\n(Bit)":
                self.LengthCol = col[3].column
            elif col[3].value == "Cycle Time":
                self.CycleCol = col[3].column
            elif col[3].value == "Sender":
                self.SenderCol = col[3].column
            elif col[3].value == "Signal Name":
                self.SigCol = col[3].column
            elif col[2].value == "Variable Name":
                self.VarNameCol = col[2].column
            elif col[2].value == "Variable Type":
                self.TypeCol = col[2].column
            elif col[2].value == "Variable 초기값":
                self.InitCol = col[2].column
            elif col[2].value == "Time-out Flag":
                self.TimoutCol = col[2].column
            elif col[2].value == "Invalid Flag":
                self.InvalCol = col[2].column
            elif col[2].value == "Invalid Flag Description":
                self.E2ECol = col[2].column
            elif col[3].value == "Factor":
                self.Factor = col[3].column
            elif col[3].value == "Offset":
                self.Offset = col[3].column

    def Ind_variable(self, sourcefile, headerfile):
        MessageInd = list()
        declare = "/*==========Message Notification variable===============*/\n"
        define = "/*==========Message Notification variable===============*/\n"
        for i in range(5, self.max_row):
            MessageInd.append(self.wb.cell(row=i, column=self.MsgCol).value) # add to massageNanme within 'MessageInd' list

        #remove duplicated factor through 'set'
        MessageInd = set(MessageInd)
        # it should be changed list type, which is able to access
        MessageInd = list(MessageInd)

        for i in MessageInd:
            declare += "extern uint8 ubS_F_{0}_Ind;\n".format(i)
            define += "uint8 ubS_F_{0}_Ind = STD_OFF;\n".format(i)

        declare += "/*=======================================================*/\n"
        define += "/*=========================================================*/\n"

        with open(sourcefile, 'a') as f:
            f.write(define)

        with open(headerfile, 'a') as f:
            f.write(declare)

    def TimeOutInd_variable(self, sourcefile, headerfile):
        TimeoutInd = list()
        TimeoutVarInd = list()
        declare = "/*==========Message Timeout Notification variable===============*/\n"
        define = "/*==========Message Timeout Notification variable===============*/\n"

        for i in range(5, self.max_row):
            TimeoutInd.append(self.wb.cell(row=i, column=self.TimoutCol).value) # add to Timeout Flag within 'TimeoutInd' list

        for j in TimeoutInd:
            if type(j) is str: # it is impossible to use 'regular expression(re)', if not str type
                m = re.match('Rxmissing.+', j)  # Rxmissing+any charactor+onemore
                if m:
                    TimeoutVarInd.append(j)
                else:
                    pass
            else :
                pass

        for i in TimeoutVarInd:
            declare += "extern uint8 {0};\n".format(i)
            define += "uint8 {0} = STD_OFF;\n".format(i)

        declare += "/*=======================================================*/\n"
        define += "/*=========================================================*/\n"

        with open(sourcefile, 'a') as f:
            f.write(define)

        with open(headerfile, 'a') as f:
            f.write(declare)

    def Invalid_variable(self, sourcefile, headerfile):
        Invalid = list()
        InvalidVar = list()
        declare = "/*==========E2E Message Invalid Notification variable===============*/\n"
        define = "/*==========E2E Message Invalid Notification variable===============*/\n"

        for i in range(5, self.max_row):
            Invalid.append(self.wb.cell(row=i, column=self.InvalCol).value) # add to Timeout Flag within 'TimeoutInd' list

        for j in Invalid:
            if type(j) is str: # it is impossible to use 'regular expression(re)', if not str type
                m = re.match('IVD.+', j)  # IVD +any charactor+onemore
                if m:
                    InvalidVar.append(j)
                else:
                    pass
            else:
                pass

        for i in InvalidVar:
            declare += "extern uint8 {0};\n".format(i)
            define += "uint8 {0} = STD_OFF;\n".format(i)

        declare += "/*=======================================================*/\n"
        define += "/*=========================================================*/\n"

        with open(sourcefile, 'a') as f:
            f.write(define)

        with open(headerfile, 'a') as f:
            f.write(declare)

    def Signal_variable(self, sourcefile, headerfile):
        CCANFD_Rx_list = list()
        PCANFD_Rx_list = list()
        GCANFD_Rx_list = list()
        CCANFD_Tx_list = list()
        PCANFD_Tx_list = list()
        GCANFD_Tx_list = list()
        CPGCANFD_Tx_list = list()

        for i in range(5, self.max_row):
            CANChannel = self.wb.cell(row=i, column=self.CANChannelCol).value
            RxTx = self.wb.cell(row=i, column=self.DirCol).value
            init = self.wb.cell(row=i, column=self.InitCol).value
            Type = self.wb.cell(row=i, column=self.TypeCol).value
            VarName = self.wb.cell(row=i, column=self.VarNameCol).value
            Length = self.wb.cell(row=i, column=self.LengthCol).value
            if CANChannel == "C-CANFD" and RxTx == "Rx":
                CCANFD_Rx_list.append([Type, VarName, init, int(Length)])
            elif CANChannel == "P-CANFD" and RxTx == "Rx":
                PCANFD_Rx_list.append([Type, VarName, init, int(Length)])
            elif CANChannel == "G-CANFD" and RxTx == "Rx":
                GCANFD_Rx_list.append([Type, VarName, init, int(Length)])
            elif CANChannel == "C-CANFD" and RxTx == "Tx":
                CCANFD_Tx_list.append([Type, VarName, init, int(Length)])
            elif CANChannel == "P-CANFD" and RxTx == "Tx":
                PCANFD_Tx_list.append([Type, VarName, init, int(Length)])
            elif CANChannel == "G-CANFD" and RxTx == "Tx":
                GCANFD_Tx_list.append([Type, VarName, init, int(Length)])
            elif CANChannel == "C/P/G-CANFD" and RxTx == "Tx":
                CPGCANFD_Tx_list.append([Type, VarName, init, int(Length)])
            else:
                pass

        m = re.compile('var_cdcu.+')  # var_cdcu + any charactor+onemore
        # one text make after getting every CAN channel variable
        CCANRxVarH = "/*==================== C CAN-FD RX variable =========================*/\n"
        CCANRxVarC = "/*==================== C CAN-FD RX variable =========================*/\n"

        for (Type, VarName, init, length) in CCANFD_Rx_list:
            a, b = codeStyle.variable(Type, VarName, init, length)
            p=m.search(a) #blank delete
            if p:
                CCANRxVarH += a
            p = m.search(b)
            if p:
                CCANRxVarC += b

        CCANRxVarH += "/*=================================================================*/\n"
        CCANRxVarC += "/*=================================================================*/\n"

        with open(headerfile, 'a') as f:
            f.write(CCANRxVarH)

        with open(sourcefile, 'a') as f:
            f.write(CCANRxVarC)

        PCANRxVarH = "/*==================== P CAN-FD RX variable =========================*/\n"
        PCANRxVarC = "/*==================== P CAN-FD RX variable =========================*/\n"

        for (Type, VarName, init, length) in PCANFD_Rx_list:
            a, b = codeStyle.variable(Type, VarName, init, length)
            p=m.search(a)
            if p:
                PCANRxVarH += a
            p = m.search(b)
            if p:
                PCANRxVarC += b

        PCANRxVarH += "/*=================================================================*/\n"
        PCANRxVarC += "/*=================================================================*/\n"

        with open(headerfile, 'a') as f:
            f.write(PCANRxVarH)

        with open(sourcefile, 'a') as f:
            f.write(PCANRxVarC)

        GCANRxVarH = "/*==================== G CAN-FD RX variable =========================*/\n"
        GCANRxVarC = "/*==================== G CAN-FD RX variable =========================*/\n"

        for (Type, VarName, init, length) in GCANFD_Rx_list:
            a, b = codeStyle.variable(Type, VarName, init, length)
            p=m.search(a)
            if p:
                GCANRxVarH += a
            p = m.search(b)
            if p:
                GCANRxVarC += b

        GCANRxVarH += "/*=================================================================*/\n"
        GCANRxVarC += "/*=================================================================*/\n"

        with open(headerfile, 'a') as f:
            f.write(GCANRxVarH)

        with open(sourcefile, 'a') as f:
            f.write(GCANRxVarC)

        CCANTxVarH = "/*==================== C CAN-FD TX variable =========================*/\n"
        CCANTxVarC = "/*==================== C CAN-FD TX variable =========================*/\n"

        for (Type, VarName, init, length) in CCANFD_Tx_list:
            a, b = codeStyle.variable(Type, VarName, init, length)
            p=m.search(a)
            if p:
                CCANTxVarH += a
            p = m.search(b)
            if p:
                CCANTxVarC += b

        CCANTxVarH += "/*=================================================================*/\n"
        CCANTxVarC += "/*=================================================================*/\n"

        with open(headerfile, 'a') as f:
            f.write(CCANTxVarH)

        with open(sourcefile, 'a') as f:
            f.write(CCANTxVarC)

        PCANTxVarH = "/*==================== P CAN-FD TX variable =========================*/\n"
        PCANTxVarC = "/*==================== P CAN-FD TX variable =========================*/\n"

        for (Type, VarName, init, length) in PCANFD_Tx_list:
            a, b = codeStyle.variable(Type, VarName, init, length)
            p=m.search(a)
            if p:
                PCANTxVarH += a
            p = m.search(b)
            if p:
                PCANTxVarC += b

        PCANTxVarH += "/*=================================================================*/\n"
        PCANTxVarC += "/*=================================================================*/\n"

        with open(headerfile, 'a') as f:
            f.write(PCANTxVarH)

        with open(sourcefile, 'a') as f:
            f.write(PCANTxVarC)

        GCANTxVarH = "/*==================== G CAN-FD TX variable =========================*/\n"
        GCANTxVarC = "/*==================== G CAN-FD TX variable =========================*/\n"

        for (Type, VarName, init, length) in GCANFD_Tx_list:
            a, b = codeStyle.variable(Type, VarName, init, length)
            p=m.search(a)
            if p:
                GCANTxVarH += a
            p = m.search(b)
            if p:
                GCANTxVarC += b

        GCANTxVarH += "/*=================================================================*/\n"
        GCANTxVarC += "/*=================================================================*/\n"

        with open(headerfile, 'a') as f:
            f.write(GCANTxVarH)

        with open(sourcefile, 'a') as f:
            f.write(GCANTxVarC)
        CPGCANTxVarH = "/*==================== CPG CAN-FD TX variable =========================*/\n"
        CPGCANTxVarC = "/*==================== CPG CAN-FD TX variable =========================*/\n"

        for (Type, VarName, init, length) in CPGCANFD_Tx_list:
            a, b = codeStyle.variable(Type, VarName, init, length)
            p=m.search(a)
            if p:
                CPGCANTxVarH += a
            p = m.search(b)
            if p:
                CPGCANTxVarC += b

        CPGCANTxVarH += "/*=================================================================*/\n"
        CPGCANTxVarC += "/*=================================================================*/\n"

        with open(headerfile, 'a') as f:
            f.write(CPGCANTxVarH)

        with open(sourcefile, 'a') as f:
            f.write(CPGCANTxVarC)

    def CCANFD_File(self, sourcefile):
        filename = os.path.basename(sourcefile)
        filename = filename.split("_")
        Source = codeStyle.Start_State(filename[0])
        E2E_TxMessage = list()
        E2E_RxMessage = list()
        CAN_Message_ind = list()
        CAN_Message_ind_E2E = list()
        CAN_RxSingal = list()
        RxFunctionSet = ""
        RxSignalSet = ""
        RxOperSet = ""
        RTERxFunction = ""
        CAN_Tx_Message = list()
        CAN_Tx_Message_E2E = list()
        CAN_TxSingal = list()
        TxSignalSet =""
        TxVarOldSet =""
        TxFunctionSet = ""
        TxOperSet = ""
        RTETxFunction = ""
        ClusterName = "CCAN"

        m = re.compile('var_cdcu.+')
        n = re.compile('.+Crc.+')

        for i in range(5, self.max_row):
            CANChannel = self.wb.cell(row=i, column=self.CANChannelCol).value
            RxTx = self.wb.cell(row=i, column=self.DirCol).value
            if CANChannel ==  "C-CANFD" and RxTx == "Rx":
                Message = self.wb.cell(row=i, column=self.MsgCol).value
                Cycle = self.wb.cell(row=i, column=self.CycleCol).value
                Signal = self.wb.cell(row=i, column=self.SigCol).value
                Variable = self.wb.cell(row=i, column=self.VarNameCol).value
                Sender = self.wb.cell(row=i, column=self.SenderCol).value
                ### add to Offset, factor ####
                Offset = self.wb.cell(row=i, column=self.Offset).value
                Factor = self.wb.cell(row=i, column=self.Factor).value
                size = self.wb.cell(row=i, column=self.TypeCol).value

                ### Search E2E Message ###
                q = n.match(str(Signal))
                if q: #include crc signal
                    E2E_RxMessage.append(Message)

                p = m.match(str(Variable))
                if p:
                   CAN_Message_ind.append([Message, int(Cycle), Sender])
                   # CAN_RxSingal.append([Message, Signal, Variable])
                   ### add to Offset, factor ####
                   CAN_RxSingal.append([size, Message, Signal, Variable, Offset, Factor])
                else:
                    pass
            elif CANChannel ==  "C-CANFD" or CANChannel == "C/P/G-CANFD" and RxTx == "Tx":
                Message = self.wb.cell(row=i, column=self.MsgCol).value
                Signal = self.wb.cell(row=i, column=self.SigCol).value
                Variable = self.wb.cell(row=i, column=self.VarNameCol).value
                size = self.wb.cell(row=i, column=self.TypeCol).value

                ### Search E2E Message ###
                q = n.match(str(Signal))
                if q:  # include crc signal
                    E2E_TxMessage.append(Message)

                p = m.match(str(Variable))
                if p:
                    CAN_Tx_Message.append(Message)
                    CAN_TxSingal.append([Message, Signal, Variable, size])
                else:
                    pass


        print("E2E_TxMessage:", E2E_TxMessage)
        print("E2E_RxMessage:", E2E_RxMessage)

        # remove duplicated factor through 'set'
        CAN_Message_ind = set(map(tuple, CAN_Message_ind)) # remove 2 dimension list
        CAN_Tx_Message = set(CAN_Tx_Message)
        # it should be changed list type, which is able to access
        CAN_Message_ind = list(CAN_Message_ind)
        CAN_Tx_Message = list(CAN_Tx_Message)

        for msg, cycle, Sender in CAN_Message_ind:
            CAN_Message_ind_E2E.append([msg, int(cycle), Sender, 0])

        for idx, a in enumerate(CAN_Message_ind_E2E):
            for temp in E2E_RxMessage:
                if a[0] == temp:
                    CAN_Message_ind_E2E[idx][3] = 1 ## E2E Message

        for msg in CAN_Tx_Message:
            CAN_Tx_Message_E2E.append([msg, 0])

        for idx, a in enumerate(CAN_Tx_Message_E2E):
            for temp in E2E_TxMessage:
                if a[0] == temp:
                    CAN_Tx_Message_E2E[idx][1] = 1 ## E2E Message

        print("CAN_TxMessage:", CAN_Tx_Message)
        print("CAN_TxMessage length:", len(CAN_Tx_Message))
        print("CAN_TxMessage E2E:", CAN_Tx_Message_E2E)
        print("CAN_TxMessage E2E length:", len(CAN_Tx_Message_E2E))

        for Name, Cycle, Sender, E2E_Flag in CAN_Message_ind_E2E:
            Source += codeStyle.Indicate_Function(filename[0], Name, Cycle) # Message Ind, Timeout Ind function create
            for size, Message, Signal, Variable, Offset, Factor in CAN_RxSingal: ### add to Offset, factor ####
                if Message == Name:
                   if E2E_Flag == 1:
                       RxSignalSet += codeStyle.Rx_E2E_Signal(size, Variable, Message, Signal, float(Offset), Factor)  # variable receive of assigned message
                   else:
                       RxSignalSet += codeStyle.Rx_Signal(size, Variable, filename[0],ClusterName, Message, Signal, float(Offset), Factor) # variable receive of assigned message
            RxFunctionSet += codeStyle.Rx_Function(ClusterName, filename[0], Name, RxSignalSet, E2E_Flag, Sender, Cycle) # merge between function and receive variable
            RxOperSet += codeStyle.Rx_FunctionSet(filename[0], Name)
            RxSignalSet = "" #initialize
        #print("source:", Source)
        RTERxFunction = codeStyle.RTE_Rx_Function(filename[0], RxOperSet)
        Source += RxFunctionSet # add Rx signal logic

        for TxMessage, E2E_Flag in CAN_Tx_Message_E2E:
            for Message, Signal, Variable, size in CAN_TxSingal:
                if TxMessage == Message:
                   TxVarOldSet += codeStyle.variable_Old(size, Variable)
                   TxSignalSet += codeStyle.Tx_Signal(Variable, filename[0],ClusterName, Message, Signal, E2E_Flag)
            TxFunctionSet += codeStyle.Tx_Function(filename[0],ClusterName, TxMessage,TxVarOldSet,TxSignalSet, E2E_Flag)
            TxOperSet += codeStyle.Tx_FunctionSet(filename[0], TxMessage)
            TxVarOldSet = "" #initialize
            TxSignalSet = "" #initialize

        Source += TxFunctionSet  # add Tx signal logic

        RTETxFunction = codeStyle.RTE_Tx_Function(filename[0], TxOperSet)

        Source += RTERxFunction # add Rx RTE function
        Source += RTETxFunction # add Tx RTE function

        with open(sourcefile, 'a') as f:
            f.write(Source)

    def PCANFD_File(self, sourcefile):
        filename = os.path.basename(sourcefile)
        filename = filename.split("_")
        Source = codeStyle.Start_State(filename[0])
        E2E_TxMessage = list()
        E2E_RxMessage = list()
        CAN_Message_ind = list()
        CAN_Message_ind_E2E = list()
        CAN_RxSingal = list()
        RxFunctionSet = ""
        RxSignalSet = ""
        RxOperSet = ""
        RTERxFunction = ""
        CAN_Tx_Message = list()
        CAN_Tx_Message_E2E = list()
        CAN_TxSingal = list()
        TxSignalSet =""
        TxVarOldSet =""
        TxFunctionSet = ""
        TxOperSet = ""
        RTETxFunction = ""
        ClusterName = "PCAN"

        m = re.compile('var_cdcu.+')
        n = re.compile('.+Crc.+')

        test = list()
        test1 = list()

        for i in range(5, self.max_row):
            CANChannel = self.wb.cell(row=i, column=self.CANChannelCol).value
            RxTx = self.wb.cell(row=i, column=self.DirCol).value
            if CANChannel ==  "P-CANFD" and RxTx == "Rx":
                Message = self.wb.cell(row=i, column=self.MsgCol).value
                Cycle = self.wb.cell(row=i, column=self.CycleCol).value
                Signal = self.wb.cell(row=i, column=self.SigCol).value
                Variable = self.wb.cell(row=i, column=self.VarNameCol).value
                Sender = self.wb.cell(row=i, column=self.SenderCol).value
                ### add to Offset, factor ####
                Offset = self.wb.cell(row=i, column=self.Offset).value
                Factor = self.wb.cell(row=i, column=self.Factor).value
                size = self.wb.cell(row=i, column=self.TypeCol).value

                ### Search E2E Message ###
                q = n.match(str(Signal))
                if q: #include crc signal
                    E2E_RxMessage.append(Message)

                p = m.match(str(Variable))
                if p:
                   CAN_Message_ind.append([Message, int(Cycle), Sender])
                   # CAN_RxSingal.append([Message, Signal, Variable])
                   ### add to Offset, factor ####
                   CAN_RxSingal.append([size, Message, Signal, Variable, Offset, Factor])
                else:
                    pass
            elif CANChannel ==  "P-CANFD" or CANChannel == "C/P/G-CANFD" and RxTx == "Tx":
                Message = self.wb.cell(row=i, column=self.MsgCol).value
                Signal = self.wb.cell(row=i, column=self.SigCol).value
                Variable = self.wb.cell(row=i, column=self.VarNameCol).value
                size = self.wb.cell(row=i, column=self.TypeCol).value

                ### Search E2E Message ###
                q = n.match(str(Signal))
                if q:  # include crc signal
                    E2E_TxMessage.append(Message)

                p = m.match(str(Variable))
                if p:
                    CAN_Tx_Message.append(Message)
                    CAN_TxSingal.append([Message, Signal, Variable, size])
                else:
                    pass


        print("E2E_TxMessage:", E2E_TxMessage)
        print("E2E_RxMessage:", E2E_RxMessage)

        # remove duplicated factor through 'set'
        CAN_Message_ind = set(map(tuple, CAN_Message_ind)) # remove 2 dimension list
        CAN_Tx_Message = set(CAN_Tx_Message)
        # it should be changed list type, which is able to access
        CAN_Message_ind = list(CAN_Message_ind)
        CAN_Tx_Message = list(CAN_Tx_Message)

        for msg, cycle, Sender in CAN_Message_ind:
            CAN_Message_ind_E2E.append([msg, int(cycle), Sender, 0])

        for idx, a in enumerate(CAN_Message_ind_E2E):
            for temp in E2E_RxMessage:
                if a[0] == temp:
                    CAN_Message_ind_E2E[idx][3] = 1 ## E2E Message

        for msg in CAN_Tx_Message:
            CAN_Tx_Message_E2E.append([msg, 0])

        for idx, a in enumerate(CAN_Tx_Message_E2E):
            for temp in E2E_TxMessage:
                if a[0] == temp:
                    CAN_Tx_Message_E2E[idx][1] = 1 ## E2E Message

        print("CAN_TxMessage:", CAN_Tx_Message)
        print("CAN_TxMessage length:", len(CAN_Tx_Message))
        print("CAN_TxMessage E2E:", CAN_Tx_Message_E2E)
        print("CAN_TxMessage E2E length:", len(CAN_Tx_Message_E2E))

        for Name, Cycle, Sender, E2E_Flag in CAN_Message_ind_E2E:
            Source += codeStyle.Indicate_Function(filename[0], Name, Cycle) # Message Ind, Timeout Ind function create
            for size, Message, Signal, Variable, Offset, Factor in CAN_RxSingal: ### add to Offset, factor ####
                if Message == Name:
                   if E2E_Flag == 1:
                       RxSignalSet += codeStyle.Rx_E2E_Signal(size, Variable, Message, Signal, float(Offset), Factor)  # variable receive of assigned message
                   else:
                       RxSignalSet += codeStyle.Rx_Signal(size, Variable, filename[0],ClusterName, Message, Signal, float(Offset), Factor) # variable receive of assigned message
            RxFunctionSet += codeStyle.Rx_Function(ClusterName, filename[0], Name, RxSignalSet, E2E_Flag, Sender, Cycle) # merge between function and receive variable
            RxOperSet += codeStyle.Rx_FunctionSet(filename[0], Name)
            RxSignalSet = "" #initialize
        #print("source:", Source)
        RTERxFunction = codeStyle.RTE_Rx_Function(filename[0], RxOperSet)
        Source += RxFunctionSet # add Rx signal logic

        for TxMessage, E2E_Flag in CAN_Tx_Message_E2E:
            for Message, Signal, Variable, size in CAN_TxSingal:
                if TxMessage == Message:
                   TxVarOldSet += codeStyle.variable_Old(size, Variable)
                   TxSignalSet += codeStyle.Tx_Signal(Variable, filename[0],ClusterName, Message, Signal, E2E_Flag)
            TxFunctionSet += codeStyle.Tx_Function(filename[0],ClusterName, TxMessage,TxVarOldSet,TxSignalSet, E2E_Flag)
            TxOperSet += codeStyle.Tx_FunctionSet(filename[0], TxMessage)
            TxVarOldSet = "" #initialize
            TxSignalSet = "" #initialize

        Source += TxFunctionSet  # add Tx signal logic

        RTETxFunction = codeStyle.RTE_Tx_Function(filename[0], TxOperSet)

        Source += RTERxFunction # add Rx RTE function
        Source += RTETxFunction # add Tx RTE function

        with open(sourcefile, 'a') as f:
            f.write(Source)

    def GCANFD_File(self, sourcefile):
        filename = os.path.basename(sourcefile)
        filename = filename.split("_")
        Source = codeStyle.Start_State(filename[0])
        E2E_TxMessage = list()
        E2E_RxMessage = list()
        CAN_Message_ind = list()
        CAN_Message_ind_E2E = list()
        CAN_RxSingal = list()
        RxFunctionSet = ""
        RxSignalSet = ""
        RxOperSet = ""
        RTERxFunction = ""
        CAN_Tx_Message = list()
        CAN_Tx_Message_E2E = list()
        CAN_TxSingal = list()
        TxSignalSet =""
        TxVarOldSet =""
        TxFunctionSet = ""
        TxOperSet = ""
        RTETxFunction = ""
        ClusterName = "GCAN"

        m = re.compile('var_cdcu.+')
        n = re.compile('.+Crc.+')

        for i in range(5, self.max_row):
            CANChannel = self.wb.cell(row=i, column=self.CANChannelCol).value
            RxTx = self.wb.cell(row=i, column=self.DirCol).value
            if CANChannel ==  "G-CANFD" and RxTx == "Rx":
                Message = self.wb.cell(row=i, column=self.MsgCol).value
                Cycle = self.wb.cell(row=i, column=self.CycleCol).value
                Signal = self.wb.cell(row=i, column=self.SigCol).value
                Variable = self.wb.cell(row=i, column=self.VarNameCol).value
                Sender = self.wb.cell(row=i, column=self.SenderCol).value
                ### add to Offset, factor ####
                Offset = self.wb.cell(row=i, column=self.Offset).value
                Factor = self.wb.cell(row=i, column=self.Factor).value
                size = self.wb.cell(row=i, column=self.TypeCol).value

                ### Search E2E Message ###
                q = n.match(str(Signal))
                if q: #include crc signal
                    E2E_RxMessage.append(Message)

                p = m.match(str(Variable))
                if p:
                   CAN_Message_ind.append([Message, int(Cycle), Sender])
                   #CAN_RxSingal.append([Message, Signal, Variable])
                   ### add to Offset, factor ####
                   CAN_RxSingal.append([size, Message, Signal, Variable, Offset, Factor])
                else:
                    pass
            elif CANChannel ==  "G-CANFD" or CANChannel == "C/P/G-CANFD" and RxTx == "Tx":
                Message = self.wb.cell(row=i, column=self.MsgCol).value
                Signal = self.wb.cell(row=i, column=self.SigCol).value
                Variable = self.wb.cell(row=i, column=self.VarNameCol).value
                size = self.wb.cell(row=i, column=self.TypeCol).value

                ### Search E2E Message ###
                q = n.match(str(Signal))
                if q:  # include crc signal
                    E2E_TxMessage.append(Message)

                p = m.match(str(Variable))
                if p:
                    CAN_Tx_Message.append(Message)
                    CAN_TxSingal.append([Message, Signal, Variable, size])
                else:
                    pass


        print("E2E_TxMessage:", E2E_TxMessage)
        print("E2E_RxMessage:", E2E_RxMessage)

        # remove duplicated factor through 'set'
        CAN_Message_ind = set(map(tuple, CAN_Message_ind)) # remove 2 dimension list
        CAN_Tx_Message = set(CAN_Tx_Message)
        # it should be changed list type, which is able to access
        CAN_Message_ind = list(CAN_Message_ind)
        CAN_Tx_Message = list(CAN_Tx_Message)

        for msg, cycle, Sender in CAN_Message_ind:
            CAN_Message_ind_E2E.append([msg, int(cycle), Sender, 0])

        for idx, a in enumerate(CAN_Message_ind_E2E):
            for temp in E2E_RxMessage:
                if a[0] == temp:
                    CAN_Message_ind_E2E[idx][3] = 1 ## E2E Message

        for msg in CAN_Tx_Message:
            CAN_Tx_Message_E2E.append([msg, 0])

        for idx, a in enumerate(CAN_Tx_Message_E2E):
            for temp in E2E_TxMessage:
                if a[0] == temp:
                    CAN_Tx_Message_E2E[idx][1] = 1 ## E2E Message

        print("CAN_TxMessage:", CAN_Tx_Message)
        print("CAN_TxMessage length:", len(CAN_Tx_Message))
        print("CAN_TxMessage E2E:", CAN_Tx_Message_E2E)
        print("CAN_TxMessage E2E length:", len(CAN_Tx_Message_E2E))

        for Name, Cycle, Sender, E2E_Flag in CAN_Message_ind_E2E:
            Source += codeStyle.Indicate_Function(filename[0], Name, Cycle) # Message Ind, Timeout Ind function create
            for size, Message, Signal, Variable, Offset, Factor in CAN_RxSingal: ### add to Offset, factor ####
                if Message == Name:
                   if E2E_Flag == 1:
                       RxSignalSet += codeStyle.Rx_E2E_Signal(size, Variable, Message, Signal, float(Offset), Factor)  # variable receive of assigned message
                   else:
                       RxSignalSet += codeStyle.Rx_Signal(size, Variable, filename[0],ClusterName, Message, Signal, float(Offset), Factor) # variable receive of assigned message
            RxFunctionSet += codeStyle.Rx_Function(ClusterName, filename[0], Name, RxSignalSet, E2E_Flag, Sender, Cycle) # merge between function and receive variable
            RxOperSet += codeStyle.Rx_FunctionSet(filename[0], Name)
            RxSignalSet = "" #initialize
        #print("source:", Source)
        RTERxFunction = codeStyle.RTE_Rx_Function(filename[0], RxOperSet)
        Source += RxFunctionSet # add Rx signal logic

        for TxMessage, E2E_Flag in CAN_Tx_Message_E2E:
            for Message, Signal, Variable, size in CAN_TxSingal:
                if TxMessage == Message:
                   TxVarOldSet += codeStyle.variable_Old(size, Variable)
                   TxSignalSet += codeStyle.Tx_Signal(Variable, filename[0],ClusterName, Message, Signal, E2E_Flag)
            TxFunctionSet += codeStyle.Tx_Function(filename[0],ClusterName, TxMessage,TxVarOldSet,TxSignalSet, E2E_Flag)
            TxOperSet += codeStyle.Tx_FunctionSet(filename[0], TxMessage)
            TxVarOldSet = "" #initialize
            TxSignalSet = "" #initialize

        Source += TxFunctionSet  # add Tx signal logic

        RTETxFunction = codeStyle.RTE_Tx_Function(filename[0], TxOperSet)

        Source += RTERxFunction # add Rx RTE function
        Source += RTETxFunction # add Tx RTE function

        with open(sourcefile, 'a') as f:
            f.write(Source)

    def TxMon_Gen(self, sourcefile):
        head = '''/*** Tx Test Code ***/
#include "interface.h"
#include "APP_Variable_KSC.h"
#include "TxMon.h"

uint8 RxTestCase = 0U;
extern void s_CAN_Rx_Mon_Test(void)
{
    if(RxTestCase==1)
    {\n'''
        end = "}"
        TxMonSig = list()
        RxMsg = list()
        RxSig = list()
        RxSignalSet = ""
        RxSigCode = ""
        Source = ""

        m = re.compile('var_cdcu_MON_.+')
        n = re.compile('var_cdcu.+')
        for i in range(5, self.max_row):
            RxTx = self.wb.cell(row=i, column=self.DirCol).value

            if RxTx == "Rx":
               Message = self.wb.cell(row=i, column=self.MsgCol).value
               Variable = self.wb.cell(row=i, column=self.VarNameCol).value
               size = self.wb.cell(row=i, column=self.TypeCol).value
               RxMsg.append(Message)

               q = n.match(str(Variable))  ## except for blank
               if q:
                    RxSig.append([size, Message, Variable])
            elif RxTx == "Tx":
                Variable = self.wb.cell(row=i, column=self.VarNameCol).value

                p = m.match(str(Variable))
                if p:
                    TxMonSig.append(Variable)

        txmonlen = len(TxMonSig)

        RxMsg = set(RxMsg)
        txmonidx = 0
        flag = 0
        cnt = 1

        for RxMessage in RxMsg:
            #for size, Message, Signal in RxSig:
            ### complete for the future ######
            for idx, tmp in enumerate(RxSig):
                #if(txmonlen < idx):
                #    mul = idx//txmonlen ## example 5//3 = 1, 5/3 = 1.6666 , 5%3 = 2
                #    txmonidx = (idx - (txmonlen*mul))-1
                #else:
                #    txmonidx = idx-1
                #################################
                if RxMessage == tmp[1]:
                   if tmp[0] == "float":
                      # index initial
                      if (txmonidx+1) >= txmonlen:
                          txmonidx = 0

                      RxSignalSet += "\t\t" + str(TxMonSig[txmonidx]) + "=(uint16)(((uint32)(" + str(tmp[2]) + ")&0xffff0000)>>16);\n" + "\t\t" + str(TxMonSig[txmonidx+1]) + "=(uint16)((uint32)(" + str(tmp[2]) + ")&0x0000ffff);\n"
                      txmonidx = txmonidx + 2

                      ## else if condition
                      if (txmonidx+1) >= txmonlen:
                          flag = 1
                          cnt = cnt+1
                      #RxSignalSet += "= (uint16)(((uint32)(" + str(Signal) + ")&0xffff0000)>>16);\n" + "= (uint16)((uint32)(" + str(Signal) + ")&0x0000ffff);\n"
                   else:
                       if txmonidx >= txmonlen:
                          txmonidx = 0
                          flag = 1
                          cnt = cnt+1
                       RxSignalSet += "\t\t" + str(TxMonSig[txmonidx]) + "=" + str(tmp[2]) + ";\n"
                       txmonidx = txmonidx + 1
                       ## else if condition
                       #if txmonidx == txmonlen:
                       #   cnt = cnt + 1
                       #   flag = 1
                       #RxSignalSet += "= " + str(Signal) + ";\n"

            RxSigCode += codeStyle.Tx_MonSig(RxMessage, RxSignalSet, flag, cnt)
            flag = 0
            RxSignalSet = ""
        Source += head
        Source += RxSigCode
        Source += end
        with open(sourcefile, 'w') as f:
            f.write(Source)

class EEpCodeGenerate:
    def __init__(self, excelpath):
        ReadExcel = openpyxl.load_workbook(excelpath, data_only=True) # pure data read except calculator of excel cell
        self.wb = ReadExcel['EEPROM Req_v2.0']
        self.max_row = self.wb.max_row
        self.max_column = self.wb.max_column

        for col in self.wb.columns: # get excel column address
            #get column number which need for code
            if col[5].value == "변수명":
                self.VarNameCol = col[5].column
            elif col[5].value == "Data Type":
                self.TypeCol = col[5].column
            elif col[5].value == "초기값\n(최초 Write 전)":
                self.InitCol = col[5].column
            elif col[5].value == "크기\n(Byte)":
                self.length = col[5].column
            elif col[5].value == "고장 정보":
                self.InvCol = col[5].column
            elif col[5].value == "Min.":
                self.MinCol = col[5].column
            elif col[5].value == "Max.":
                self.MaxCol = col[5].column

    def Eep_variable(self):
        EepVariable_list = list()
        InvVariable_list = list()
        EepVarH = "/*==========EEPROM variable===============*/\n"
        EepVarC = "/*==========EEPROM variable===============*/\n"

        for i in range(7, self.max_row):
            VarName = self.wb.cell(row=i, column=self.VarNameCol).value
            Type = self.wb.cell(row=i, column=self.TypeCol).value
            IvdVar = self.wb.cell(row=i, column=self.InvCol).value
            length = self.wb.cell(row=i, column=self.length).value
            EepVariable_list.append([Type, VarName, length])
            InvVariable_list.append(IvdVar)

        for (Type, VarName, length) in EepVariable_list:
            if Type is not None:
                a, b = codeStyle.EEp_variable(Type, VarName, length)
                EepVarH += a
                EepVarC += b

        for Inv in InvVariable_list:
            if Inv is not None:
                EepVarH += "\nextern uint8 {var};".format(var=Inv)
                EepVarC += "\nuint8 {var};".format(var=Inv)

        EepVarH += "\n/*=================================================================*/\n"
        EepVarC += "\n/*=================================================================*/\n"

        return EepVarH, EepVarC

    def NvM_File(self, sourcefile):
        Eep_list = list()
        Eep_Val_list = list()

        Source = "/******************** JOB Finished function********************************/\n"

        for i in range(7, self.max_row):
            VarName = self.wb.cell(row=i, column=self.VarNameCol).value
            Type = self.wb.cell(row=i, column=self.TypeCol).value
            Init = self.wb.cell(row=i, column=self.InitCol).value
            Max = self.wb.cell(row=i, column=self.MaxCol).value
            Min = self.wb.cell(row=i, column=self.MinCol).value

            if Type is not None:
                Eep_list.append(VarName)
                Eep_Val_list.append([VarName,Init, Max, Min])

        print("Eep_Val_list:", Eep_Val_list)

        for NvM in Eep_list:
            Source += codeStyle.JobFinished(NvM)
            Source += codeStyle.JobFinished_Mirr(NvM)

        Source += codeStyle.Init_Header()

        for VarName,Init, Max, Min in Eep_Val_list:
            Source += codeStyle.Init_Nvm(VarName, Init, Max, Min)

        Source += "\n}"

        with open(sourcefile, 'w', encoding='UTF-8') as f: # It need to encode to UTF-8 because of including cp949 at Maxvalue
            f.write(Source)